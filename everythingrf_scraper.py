#!/usr/bin/env python3
"""
everythingrf_scraper.py
=======================
Pulls the everythingRF company directory (https://www.everythingrf.com/companies,
~2,598 companies) into a structured Excel workbook so it can be sorted / filtered /
pivoted for market analysis.

WHY THIS RUNS LOCALLY (not inside claude.ai):
  The claude.ai bash sandbox has no outbound internet, so the bulk fetch can't run
  there. Run this on your laptop or via Claude Code, where Python can reach the web.

WHAT IT PRODUCES:
  everythingrf_companies.xlsx
    - Sheet "Companies"          : one row per company (the master table)
    - Sheet "Company_Categories" : one row per company x product category (for pivots)
  everythingrf_companies.csv     : same master table, flat
  cache/<id>.html                : raw HTML per company (so re-runs don't re-fetch)
  company_index.csv              : enumerated id/name/country/url checkpoint

TWO MODES (set MODE below):
  "list"  : fast. Name, country, profile URL only. ~52 page loads, ~1-2 min.
  "full"  : adds website, address, phone, description, product categories +
            product counts, certifications, LinkedIn. ~2,600 page loads.
            At DELAY=1.0s that's ~45-50 min. Resumable — safe to stop/restart.

RESPONSIBLE USE:
  - everythingRF is a free, ad-supported directory. Keep DELAY polite (>=1s),
    run it once, and don't hammer the server. The script is single-threaded on purpose.
  - Company names / countries / addresses are factual business data; a compiled
    directory may still carry Terms-of-Service restrictions. For a recurring need,
    the clean path is to ask everythingRF for a data export or license.
    This script is for a one-off pull.
  - Put a real contact in CONTACT below — it's good etiquette and helps the site
    owner reach you instead of just blocking the traffic.

INSTALL:
  pip install requests beautifulsoup4 lxml pandas openpyxl

RUN:
  python everythingrf_scraper.py              # uses MODE below
  python everythingrf_scraper.py --mode list  # override
  python everythingrf_scraper.py --mode full
  python everythingrf_scraper.py --mode full --country "United States"

NOTE ON THE DETAIL PARSER:
  The list-page parser (name/country/id/url) is reliable. The detail-page selectors
  (address, phone, categories, product counts, certs) are best-effort guesses against
  the live HTML and are isolated in parse_detail(). If a field comes back blank, open
  one cached cache/<id>.html, adjust the selector in parse_detail(), and re-run — the
  cache means you won't re-download anything.
"""

import argparse
import csv
import json
import os
import re
import sys
import time
from datetime import datetime, timezone

import requests
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ----------------------------- CONFIG -----------------------------
MODE = "full"                 # "list" or "full"
DELAY = 1.0                   # seconds between requests (be polite)
OUTPUT_DIR = "."              # where files are written
CACHE_DIR = "cache"           # per-company raw HTML cache (enables resume)
CONTACT = "your.email@example.com"   # <-- put a real address here
COUNTRY_FILTER = None         # e.g. "United States" to restrict, else None
STARTSWITH_FILTER = None      # e.g. "a" for A-only, else None (all)
MAX_PAGES = 200               # safety cap on list pagination
# ------------------------------------------------------------------

BASE = "https://www.everythingrf.com"
LIST_URL = BASE + "/companies-category=na-country={country}-startswith={sw}-page={page}"

# Countries from the directory's own filter dropdown — used to reliably pull the
# country token out of each list row and detail header. Longest first so multi-word
# names match before their substrings.
COUNTRIES = sorted([
    "Andorra", "Australia", "Austria", "Belgium", "Bermuda", "Brazil", "Canada",
    "China", "Czech Republic", "Denmark", "England", "Estonia", "Finland", "France",
    "Germany", "Greece", "Hong Kong", "Hungary", "India", "Ireland", "Israel",
    "Italy", "Japan", "Korea - South", "Korea, Republic of", "Latvia", "Lithuania",
    "Luxembourg", "Malaysia", "Mexico", "Netherlands", "New Zealand", "Norway",
    "Poland", "Romania", "Russia", "Serbia", "Singapore", "Slovakia", "South Africa",
    "South Korea", "Spain", "Sweden", "Switzerland", "Taiwan", "Thailand", "Turkey",
    "Ukraine", "United Kingdom", "United States", "Viet Nam", "Vietnam",
], key=len, reverse=True)

COMPANY_HREF = re.compile(r"/companies/(\d+)/([^/?#\"']+)")
PHONE_RE = re.compile(r"[\+\(]?\d[\d\s\-\(\)\.]{6,}\d")
COUNT_RE = re.compile(r"\(([\d,]+)\)")


def make_session():
    s = requests.Session()
    retry = Retry(total=5, backoff_factor=1.5,
                  status_forcelist=[429, 500, 502, 503, 504],
                  allowed_methods=["GET"])
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.headers.update({
        "User-Agent": (f"Mozilla/5.0 (compatible; DirectoryResearchBot/1.0; +{CONTACT}) "
                       "company-directory-research"),
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "en-US,en;q=0.9",
    })
    return s


def get(session, url):
    r = session.get(url, timeout=30)
    r.raise_for_status()
    time.sleep(DELAY)
    return r.text


def match_country(text):
    """Return the first known country found as a trailing token in text, else ''."""
    t = " ".join(text.split())
    for c in COUNTRIES:
        if t.endswith(c) or f" {c} " in f" {t} ":
            return "South Korea" if c in ("Korea - South", "Korea, Republic of") else c
    return ""


# --------------------------- ENUMERATION ---------------------------
def parse_list_page(html):
    """Return list of dicts {id, name, country, url} from one directory page."""
    soup = BeautifulSoup(html, "lxml")
    seen, out = set(), []
    for a in soup.find_all("a", href=COMPANY_HREF):
        m = COMPANY_HREF.search(a["href"])
        if not m:
            continue
        cid, slug = m.group(1), m.group(2)
        if cid in seen or slug == "add-your-company":
            continue
        name = a.get_text(strip=True)
        if not name:
            continue
        # Country sits in the same row as the link; grab the row's text and parse it.
        row = a.find_parent(["li", "tr", "div"]) or a
        row_text = re.sub(r"\(.*?\)", "", row.get_text(" ", strip=True))  # drop "(Type)"
        country = match_country(row_text)
        seen.add(cid)
        out.append({"id": cid, "name": name,
                    "url": f"{BASE}/companies/{cid}/{slug}", "country": country})
    return out


def enumerate_companies(session):
    sw = STARTSWITH_FILTER or "na"
    country = "na"  # the list URL's country slug; we filter by parsed country instead
    index, prev_ids = {}, None
    for page in range(1, MAX_PAGES + 1):
        url = LIST_URL.format(country=country, sw=sw, page=page)
        try:
            html = get(session, url)
        except Exception as e:
            print(f"  ! page {page} failed: {e}")
            break
        rows = parse_list_page(html)
        ids = {r["id"] for r in rows}
        if not rows or ids == prev_ids:   # empty page or same page repeating = end
            break
        prev_ids = ids
        for r in rows:
            index[r["id"]] = r
        print(f"  page {page}: +{len(rows)} (total {len(index)})")
    companies = list(index.values())
    if COUNTRY_FILTER:
        companies = [c for c in companies if c["country"] == COUNTRY_FILTER]
    return companies


# ----------------------------- DETAIL ------------------------------
def strip_tracking(href):
    return re.sub(r"[?&]utm_[^=]+=[^&]+", "", href).rstrip("?&")


def parse_detail(html, base_record):
    """Best-effort detail-page parse. Selectors isolated here for easy tuning."""
    soup = BeautifulSoup(html, "lxml")
    rec = dict(base_record)
    rec.update({"city": "", "state_region": "", "address": "", "phone": "",
                "website": "", "linkedin": "", "certifications": "",
                "num_categories": 0, "total_products": 0,
                "categories_supported": "", "description": ""})

    # Name (breadcrumb h1 is the most stable anchor)
    h1 = soup.find("h1")
    if h1 and h1.get_text(strip=True):
        rec["name"] = h1.get_text(strip=True)

    # Header contact block: classify the short lines that sit just under the h1.
    header_lines = []
    if h1:
        for el in h1.find_all_next(["li", "p", "span", "div"], limit=40):
            txt = el.get_text(" ", strip=True)
            if not txt or len(txt) > 160:
                continue
            if txt in header_lines:
                continue
            header_lines.append(txt)
            if "Visit Website" in txt or "Contact Company" in txt:
                break
    for line in header_lines:
        if not rec["country"]:
            c = match_country(line)
            if c and len(line) < 40:
                rec["country"] = c
        if not rec["phone"] and PHONE_RE.fullmatch(line.strip()):
            rec["phone"] = line.strip()
        if not rec["address"] and "," in line and any(ch.isdigit() for ch in line) \
                and not PHONE_RE.fullmatch(line.strip()) and len(line) > 12:
            rec["address"] = line.strip()

    # City / state from address (best-effort; tuned for "City, ST ZIP, Country")
    if rec["address"]:
        parts = [p.strip() for p in rec["address"].split(",")]
        if parts and match_country(parts[-1]):
            parts = parts[:-1]
        if len(parts) >= 2:
            rec["city"] = parts[-2]
            m = re.match(r"([A-Za-z\.\s]+)\s+\d", parts[-1])
            rec["state_region"] = (m.group(1).strip() if m else parts[-1])

    # Website ("Visit Website" link)
    wa = soup.find("a", string=re.compile(r"Visit Website", re.I))
    if not wa:
        wa = soup.find("a", href=re.compile(r"utm_source=everythingrf"))
    if wa and wa.get("href"):
        rec["website"] = strip_tracking(wa["href"])

    # LinkedIn
    li = soup.find("a", href=re.compile(r"linkedin\.com/company", re.I))
    if li:
        rec["linkedin"] = li["href"]

    # Description: the first substantial paragraph in the main content.
    for p in soup.find_all("p"):
        t = p.get_text(" ", strip=True)
        if len(t) > 120:
            rec["description"] = t
            break

    # Categories Supported: links to /directory/ under that heading.
    cats = []
    for a in soup.find_all("a", href=re.compile(r"/directory/")):
        name = a.get_text(strip=True)
        if name and "filters?" not in a["href"] and name not in cats:
            cats.append(name)
    rec["categories_supported"] = "; ".join(cats)
    rec["num_categories"] = len(cats)

    # Products Listed with counts: /search/ links whose tail holds "(1,234)".
    prod_rows, total = [], 0
    for a in soup.find_all("a", href=re.compile(r"/search/.*manuid=")):
        cat = a.get_text(strip=True)
        tail = (a.next_sibling or "")
        tail = tail if isinstance(tail, str) else tail.get_text(" ", strip=True)
        m = COUNT_RE.search(tail or "")
        n = int(m.group(1).replace(",", "")) if m else 0
        if cat:
            prod_rows.append((cat, n))
            total += n
    rec["total_products"] = total
    rec["_product_rows"] = prod_rows  # consumed by the Company_Categories sheet

    # Certifications: short lines under a "Certifications" heading.
    head = soup.find(lambda t: t.name in ("h2", "h3", "h4", "strong")
                     and "Certification" in t.get_text())
    certs = []
    if head:
        for el in head.find_all_next(limit=12):
            t = el.get_text(" ", strip=True)
            if t and len(t) < 40 and re.search(r"ISO|AS\d|IATF|MIL|RoHS|ITAR", t, re.I):
                if t not in certs:
                    certs.append(t)
    rec["certifications"] = "; ".join(certs)
    return rec


# ----------------------------- OUTPUT ------------------------------
def write_excel(records, path):
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment

    cols = ["id", "name", "country", "city", "state_region", "address", "phone",
            "website", "linkedin", "certifications", "num_categories",
            "total_products", "categories_supported", "description", "url"]
    nice = ["everythingRF_ID", "Company", "Country", "City", "State/Region",
            "Address", "Phone", "Website", "LinkedIn", "Certifications",
            "Num_Categories", "Total_Products", "Categories_Supported",
            "Description", "Profile_URL"]
    df = pd.DataFrame([{c: r.get(c, "") for c in cols} for r in records])
    df.columns = nice
    df.insert(len(df.columns), "Scraped_At",
              datetime.now(timezone.utc).strftime("%Y-%m-%d"))

    long_rows = []
    for r in records:
        for cat, n in r.get("_product_rows", []):
            long_rows.append({"everythingRF_ID": r["id"], "Company": r["name"],
                              "Country": r.get("country", ""),
                              "Category": cat, "Product_Count": n})
    long_df = pd.DataFrame(long_rows,
                           columns=["everythingRF_ID", "Company", "Country",
                                    "Category", "Product_Count"])

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="Companies", index=False)
        long_df.to_excel(xw, sheet_name="Company_Categories", index=False)
        for name in ("Companies", "Company_Categories"):
            ws = xw.sheets[name]
            for cell in ws[1]:
                cell.font = Font(name="Arial", bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="23283C")
                cell.alignment = Alignment(horizontal="left")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                width = min(max(len(str(c.value)) for c in col if c.value) + 2, 60)
                ws.column_dimensions[col[0].column_letter].width = width
    df.to_csv(path.replace(".xlsx", ".csv"), index=False, quoting=csv.QUOTE_ALL)


# ------------------------------ MAIN -------------------------------
def main():
    global MODE, COUNTRY_FILTER
    ap = argparse.ArgumentParser()
    ap.add_argument("--mode", choices=["list", "full"], default=MODE)
    ap.add_argument("--country", default=COUNTRY_FILTER)
    args = ap.parse_args()
    MODE, COUNTRY_FILTER = args.mode, args.country

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(CACHE_DIR, exist_ok=True)
    session = make_session()

    print(f"Enumerating directory (mode={MODE}, country={COUNTRY_FILTER or 'all'}) ...")
    companies = enumerate_companies(session)
    print(f"Found {len(companies)} companies.")

    with open(os.path.join(OUTPUT_DIR, "company_index.csv"), "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["id", "name", "country", "url"])
        w.writeheader()
        w.writerows(companies)

    if MODE == "list":
        records = companies
    else:
        records = []
        for i, c in enumerate(companies, 1):
            cache = os.path.join(CACHE_DIR, f"{c['id']}.html")
            try:
                if os.path.exists(cache):
                    html = open(cache, encoding="utf-8").read()
                else:
                    html = get(session, c["url"])
                    open(cache, "w", encoding="utf-8").write(html)
                records.append(parse_detail(html, c))
            except Exception as e:
                print(f"  ! {c['name']} ({c['url']}) failed: {e}")
                records.append(c)
            if i % 25 == 0 or i == len(companies):
                print(f"  detail {i}/{len(companies)}")

    out = os.path.join(OUTPUT_DIR, "everythingrf_companies.xlsx")
    write_excel(records, out)
    print(f"\nDone -> {out}")
    print(f"       {out.replace('.xlsx', '.csv')}")
    print(f"       company_index.csv  ({len(companies)} rows)")


if __name__ == "__main__":
    sys.exit(main())
